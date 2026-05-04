# coding='utf-8'
import xlrd
import re
import requests
import time as t
import random
import urllib3
from multiprocessing.dummy import Pool as ThreadPool
import openpyxl
urllib3.disable_warnings()
from concurrent.futures import ThreadPoolExecutor
cookie1 = ('_T_WM=450a866f9b4e835b21f805b6b5ec6bcf; '
           'SCF=Akt48KpW9yWSlKI0WB9w2V-YyQLO4GRHuuXWT4vPRKxj1hzwt9Tb87na9i9er12sP1_lF_f_n0hieWL1MXeY0ks.;'
           ' SUB=_2A25LOBnSDeRhGeFJ7VQY9SfIyTqIHXVoNBMarDV6PUJbktB-LUzzkW1Nf07ABz09gO1rRGtZjk4w4Sx3rVJ73N9F; '
           'SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9WFalAe_pP0e7AMKw8.Nv.IS5NHD95QNS0qc1K-4ShzcWs4Dqcj-PEH81CHWeE-RxCH8SE-RBbHF18Yt; '
           'SSOLoginState=1715235201; '
           'ALF=1717827201')  # 微博的cookie
cookie2 = ('_T_WM=71c3e86f02d0c0ec9a323d960d8869dc; '
           'SCF=AjFU9q2Z7sJKxMc4gZsBdbsMRAutpmvN1ZvMXAe0Wj6mctBtfaVYKM5IWjpE9_n01RlxN_QwuHuVdZ1IAclJVvU.; '
           'SUB=_2A25LOMs4DeRhGeFJ4loS9y3NzzqIHXVoNELwrDV6PUJbktANLXmlkW1NfrTbcXxvw81RvtFPyM6mhaZEqWJk1Zbr; '
           'SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9WFuQe8YMpNiF_J.2aiM8nUO5NHD95QNS0.Re0M0eKBcWs4Dqc_Fi--Ni-2NiKyFi--fiKysi-zci--Ni-82iKn4i--Xi-zRiKy2i--fiKnRi-z7i--RiKyWi-zpi--fi-ihiKn7eKnR; '
           'SSOLoginState=1715256168; '
           'ALF=1717848168')
headers1 = {'Accept-Encoding': 'gzip, deflate, sdch',
            'Accept-Language': 'en-US,en;q=0.8',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Referer': 'https://www.baidu.com/',
            'Connection': 'keep-alive',
            'Cookie': cookie1}
headers2 = {'Accept-Encoding': 'gzip, deflate, sdch',
            'Accept-Language': 'en-US,en;q=0.8',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Referer': 'https://www.baidu.com/',
            'Connection': 'keep-alive',
            'Cookie': cookie2}

def extract(inpath, l):
    wb = openpyxl.load_workbook(inpath)
    sheet = wb.active
    numbers = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=l, max_col=l):
        for cell in row:
            numbers.append(cell.value)
    return numbers
def require(url):
    while True:
        try:
            # 随机选择 headers1 或 headers2
            headers = random.choice([headers1, headers2])
            response = requests.get(url, headers=headers, timeout=(30, 50), verify=False)
            code_1 = response.status_code
            if code_1 == 200:
                print('正常爬取中，状态码：' + str(code_1))
                t.sleep(random.randint(2, 9))
                break
            else:
                print('请求异常，重试中，状态码为：' + str(code_1))
                t.sleep(random.uniform(60, 65))
                continue
        except:
            t.sleep(random.uniform(60, 65))
            continue
    html = response.text
    return html
def body(html):
    """单个资料爬取"""
    all_info=[]
    data = re.findall('<div class="tip">基本信息</div>(.*?)<div class="tip">其他信息</div>', html, re.S)  # 取大
    # print(data)
    name_list = re.findall('<div class="c">昵称:(.*?)<br/>', str(data), re.S)  # 用户昵称
    sex_list = re.findall('<br/>性别:(.*?)<br/>', str(data), re.S)  # 性别
    region_list = re.findall('<br/>地区:(.*?)<br/>', str(data), re.S)  # 地区
    birthdate_list = re.findall('<br/>生日:(\d{4}-\d{1,2}-\d{1,2})<br/>', str(data), re.S)  # 生日
    if name_list:
        name=name_list[0]
    else:
        name="N/A"
    if sex_list:
        sex=sex_list[0]
    else:
        sex="N/A"
    if region_list:
        region=region_list[0]
    else:
        region="N/A"
    if birthdate_list:
        birthdate=birthdate_list[0]
    else:
        birthdate="N/A"
    # print(name_0)
    all_info.append([name, sex, region, birthdate])
    # all = name + sex + region + birthdate
    return all_info
def run(uid):
    # uid=int(uid)
    url = 'https://weibo.cn/' + str(uid) + '/info'
    pa=[]
    try:
        one_data = body(require(url))
        print(one_data)
        return one_data
    except:
        return pa

def save_to_excel(data, output_path):
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Write headers
    headers = ['昵称', '性别', '地区', '生日']
    sheet.append(headers)

    # Write data
    for item in data:
        sheet.append(item[0])

    # Save the workbook
    wb.save(output_path)

if __name__ == '__main__':
    uid = extract('胖猫.xlsx', 3)
    uids = []
    processed_uids = set()
    for i in uid:
        if i not in processed_uids:
            uids.append(i)
            processed_uids.add(i)
    alls_1 = []
    with ThreadPoolExecutor() as executor:
        alls_1 = executor.map(run, uids)
    alls_2 = []
    for i in alls_1:
        if i:
            alls_2.append(i)
    save_to_excel(alls_2, '用户信息.xlsx')
