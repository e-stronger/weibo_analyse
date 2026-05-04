#coding='utf-8'
import re
import requests
import time as t
import openpyxl
from multiprocessing.dummy import Pool as ThreadPool
import random
import urllib3
import xlwt
urllib3.disable_warnings()
cookie1=('_T_WM=450a866f9b4e835b21f805b6b5ec6bcf; '
        'SCF=Akt48KpW9yWSlKI0WB9w2V-YyQLO4GRHuuXWT4vPRKxj1hzwt9Tb87na9i9er12sP1_lF_f_n0hieWL1MXeY0ks.;'
        ' SUB=_2A25LOBnSDeRhGeFJ7VQY9SfIyTqIHXVoNBMarDV6PUJbktB-LUzzkW1Nf07ABz09gO1rRGtZjk4w4Sx3rVJ73N9F; '
        'SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9WFalAe_pP0e7AMKw8.Nv.IS5NHD95QNS0qc1K-4ShzcWs4Dqcj-PEH81CHWeE-RxCH8SE-RBbHF18Yt; '
        'SSOLoginState=1715235201; '
         'ALF=1717827201')#微博的cookie
cookie2=('_T_WM=71c3e86f02d0c0ec9a323d960d8869dc; '
         'SCF=AjFU9q2Z7sJKxMc4gZsBdbsMRAutpmvN1ZvMXAe0Wj6mctBtfaVYKM5IWjpE9_n01RlxN_QwuHuVdZ1IAclJVvU.; '
         'SUB=_2A25LOMs4DeRhGeFJ4loS9y3NzzqIHXVoNELwrDV6PUJbktANLXmlkW1NfrTbcXxvw81RvtFPyM6mhaZEqWJk1Zbr; '
         'SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9WFuQe8YMpNiF_J.2aiM8nUO5NHD95QNS0.Re0M0eKBcWs4Dqc_Fi--Ni-2NiKyFi--fiKysi-zci--Ni-82iKn4i--Xi-zRiKy2i--fiKnRi-z7i--RiKyWi-zpi--fi-ihiKn7eKnR; '
         'SSOLoginState=1715256168; '
         'ALF=1717848168')
headers1 = {'Accept-Encoding': 'gzip, deflate, sdch',
		   'Accept-Language': 'en-US,en;q=0.8',
		   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0',
		   'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
		   'Referer': 'https://www.baidu.com/',
		   'Connection': 'keep-alive',
		   'Cookie': cookie1}

headers2 = {'Accept-Encoding': 'gzip, deflate, sdch',
		   'Accept-Language': 'en-US,en;q=0.8',
		   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0',
		   'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
		   'Referer': 'https://www.baidu.com/',
		   'Connection': 'keep-alive',
		   'Cookie': cookie2}

def require(url):
    """获取网页源码"""
    while True:
        try:
            # 随机选择 headers1 或 headers2
            headers = random.choice([headers1, headers2])
            response = requests.get(url, headers=headers, timeout=(30,50), verify=False)
            code_1 = response.status_code
            if code_1 == 200:
                print('正常爬取中，状态码：' + str(code_1))
                t.sleep(random.randint(2,9))
                break
            else:
                print('请求异常，重试中，状态码为：' + str(code_1))
                t.sleep(random.uniform(60,65))
                continue
        except:
            t.sleep(random.uniform(60,65))
            continue
    html = response.text
    return html

def body(h_1):
    html_2 = re.findall('<div class="c" id="M_">(.*?)</div>', h_1, re.S)
    all_comments = []
    for comment_html in html_2:
        user_id_list = re.findall('<a href="/(.*?)">(.*?)</a>', comment_html, re.S)
        if user_id_list:  # 检查是否成功匹配到用户ID列表
            user_id = user_id_list[0][0]
            name = user_id_list[0][1]
        else:
            user_id = "N/A"  # 如果未成功匹配到用户ID，则将其设置为 "N/A"
            name = "N/A"
        content_list = re.findall('<span class="ctt">(.*?)</span>', comment_html, re.S)
        if content_list:  # 检查是否成功匹配到评论内容列表
            content = content_list[0]
            content = re.sub('<.*?>', '', content)
            content = content.strip()
        else:
            content = "N/A"  # 如果未成功匹配到评论内容，则将其设置为 "N/A"
        time_list = re.findall('<span class="ct">(.*?)</span>', comment_html, re.S)
        if time_list:  # 检查是否成功匹配到时间列表
            time = time_list[0]
        else:
            time = "N/A"  # 如果未成功匹配到时间，则将其设置为 "N/A"
        all_comments.append([user_id, name, content, time])
    return all_comments

def save_to_excel(data, filename):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "评论信息"
    # 写入表头
    headers = ['用户ID', '用户名', '评论内容', '时间']
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)
    # 写入数据
    for row, comment in enumerate(data, start=2):
        for col, value in enumerate(comment, start=1):
            sheet.cell(row=row, column=col, value=value)
    # 保存文件
    wb.save(filename)

def extract(inpath, l):
    """从 Excel（.xlsx）文件中提取一列数据。"""
    wb = openpyxl.load_workbook(inpath)
    sheet = wb.active
    numbers = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=l, max_col=l):
        for cell in row:
            numbers.append(cell.value)
    return numbers

def run(ids):
    b = ids[0]  # bid
    u = str(ids[1]).replace('.0', '')  # uid
    pa = []  # 空列表判定
    url = 'https://weibo.cn/comment/' + str(b) + '?uid=' + str(u)  # 一个微博的评论首页
    html = require(url)
    try:
        data_1 = body(html)
        return data_1
    except:
        return pa


if __name__ == '__main__':
    # 由于微博限制，只能爬取前五十页的
    # 里面的文件是爬取到的正文文件
    bid = extract('胖猫.xlsx', 2)  # 1是bid，2是u_id
    uid = extract('胖猫.xlsx', 1)
    ids = []  # 将bid和uid匹配并以嵌套列表形式加入ids
    for i, j in zip(bid, uid):
        ids.append([i, j])
    # 多线程
    pool = ThreadPool()
    all_comments = pool.map(run, ids)
    # 合并所有评论数据到一个列表中
    all_comments_flat = [comment for sublist in all_comments if sublist for comment in sublist]
    # 保存评论数据到 Excel 文件
    save_to_excel(all_comments_flat, '评论信息2.xlsx')