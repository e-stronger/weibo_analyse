import openpyxl
def remove_duplicates(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    sheet = wb.active

    unique_wb = openpyxl.Workbook()
    unique_sheet = unique_wb.active

    seen_rows = set()

    for row in sheet.iter_rows(values_only=True):
        row_tuple = tuple(row)

        if row_tuple not in seen_rows:
            unique_sheet.append(row)
            seen_rows.add(row_tuple)

    unique_wb.save(output_path)


# Usage example
input_file = '评论信息.xlsx'
output_file = '评论信息（去重版）.xlsx'
remove_duplicates(input_file, output_file)